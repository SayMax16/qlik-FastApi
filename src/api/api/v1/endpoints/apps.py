from fastapi import APIRouter, Depends, Path, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from src.api.services.app_service import AppService
from src.api.core.dependencies import get_app_service, verify_api_key
from src.api.core.config import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

router = APIRouter()

# IMPORTANT: More specific routes must come BEFORE generic routes
# These specific routes must be defined before the generic {table_name} route

@router.get("/apps/{app_name}/tables/factory_data/data")
async def get_factory_data(
    app_name: str = Path(..., description="Application name"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get factory data with optional filtering and variable controls.

    This endpoint retrieves data from the factory_data pivot table (object ID: Dkjpv)
    and supports filtering by Завод (factory) and Склад (warehouse) fields,
    plus setting variables for measure type and currency.

    **Examples:**

    Get all data (no filtering):
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100
    ```

    Filter by single factory:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&factory=1203
    ```

    Filter by multiple factories:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&factory=1203,1204
    ```

    Filter by warehouse:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&warehouse=A100
    ```

    Filter by OM type:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&typeOM=Type1
    ```

    Filter by factory, warehouse, and OM type with variables:
    ```
    GET /api/v1/apps/afko/tables/factory_data/data?page=1&page_size=100&factory=1203&warehouse=A100&typeOM=Type1&MeasureType=1&Currency=2
    ```

    **Response format:**
    ```json
    {
      "object_id": "Dkjpv",
      "app_id": "...",
      "app_name": "afko",
      "data": [
        {
          "Field1": "Value1",
          "Field2": "Value2",
          ...
        }
      ],
      "pagination": {
        "page": 1,
        "page_size": 100,
        "total_rows": 150,
        "total_pages": 2,
        "has_next": true,
        "has_previous": false
      }
    }
    ```
    """
    table_name = "factory_data"

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get bookmark ID for this table (applies 3-month filter first)
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # Build selections dictionary for Qlik filtering
    # NOTE: yearMonth is NOT included in selections - it will be filtered client-side
    selections = {}
    if factory:
        # Split comma-separated values
        factory_values = [f.strip() for f in factory.split(',')]
        selections['Завод'] = factory_values  # Завод is the factory field in Qlik

    if warehouse:
        # Split comma-separated values
        warehouse_values = [w.strip() for w in warehouse.split(',')]
        selections['Склад'] = warehouse_values  # Склад is the warehouse field in Qlik

    if typeOM:
        # Split comma-separated values
        typeOM_values = [t.strip() for t in typeOM.split(',')]
        selections['Тип ОМ'] = typeOM_values  # Тип ОМ is the OM type field in Qlik

    # Build filters dictionary for client-side filtering
    # yearMonth will be filtered by parsing the Дата (Date) field after fetching from Qlik
    filters = {}
    if yearMonth:
        # Parse yearMonth values for client-side filtering
        # Input format: "2024-01" or "2024.01"
        year_month_list = []
        for ym in yearMonth.split(','):
            ym = ym.strip()
            # Normalize to YYYY-MM format
            ym_normalized = ym.replace('.', '-')
            year_month_list.append(ym_normalized)
        filters['yearMonth'] = year_month_list

    # Build variables dictionary
    variables = {}
    if MeasureType:
        variables['vChooseType'] = MeasureType
    if Currency:
        variables['vChooseCur'] = Currency

    # Fetch data with bookmark applied first (filters to 3 months), then apply selections
    # yearMonth is filtered client-side by parsing the Дата field
    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=page,
        page_size=page_size,
        filters=filters,  # Client-side filtering (yearMonth by Дата field)
        selections=selections,  # Apply Qlik selections for filtering
        variables=variables,  # Apply Qlik variables
        bookmark_id=bookmark_id  # Apply bookmark FIRST to filter to 3 months
    )

    return data


@router.get("/apps/{app_name}/tables/factory_data/export")
async def export_factory_data_to_excel(
    app_name: str = Path(..., description="Application name"),
    factory: Optional[str] = Query(None, description="Filter by factory (Завод field), supports multiple values separated by comma"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (Склад field), supports multiple values separated by comma"),
    typeOM: Optional[str] = Query(None, description="Filter by OM type (Тип ОМ field), supports multiple values separated by comma"),
    yearMonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2024-01 or 2024.01), supports multiple values separated by comma"),
    MeasureType: Optional[str] = Query(None, description="Measure type (1=qty, 2=amount, 3=amount-qty)"),
    Currency: Optional[str] = Query(None, description="Currency type (1=ZUD, 2=UZS, 3=ZUDMVP)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Export factory data to Excel file.

    This endpoint exports all factory data (with applied filters) to an Excel file.
    The Excel file includes formatted headers and all data rows.

    **Examples:**

    Export all data for factory 1203:
    ```
    GET /api/v1/apps/afko/tables/factory_data/export?factory=1203
    ```

    Export with multiple filters:
    ```
    GET /api/v1/apps/afko/tables/factory_data/export?factory=1203&yearMonth=2026-03&typeOM=Отгрузка в РЦ
    ```
    """

    # Get bookmark ID for this table
    table_name = "factory_data"
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # Build selections dictionary (same as JSON endpoint)
    selections = {}
    if factory:
        factory_values = [f.strip() for f in factory.split(',')]
        selections['Завод'] = factory_values

    if warehouse:
        warehouse_values = [w.strip() for w in warehouse.split(',')]
        selections['Склад'] = warehouse_values

    if typeOM:
        typeOM_values = [t.strip() for t in typeOM.split(',')]
        selections['Тип ОМ'] = typeOM_values

    # Build filters dictionary for client-side filtering
    filters = {}
    if yearMonth:
        yearMonth_values = [ym.strip().replace('.', '-') for ym in yearMonth.split(',')]
        filters['yearMonth'] = yearMonth_values

    # Build variables dictionary
    variables = {}
    if MeasureType:
        variables['vChooseType'] = MeasureType
    if Currency:
        variables['vChooseCur'] = Currency

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=1,
        page_size=100000,  # Fetch all data
        filters=filters,
        selections=selections,
        variables=variables,
        bookmark_id=bookmark_id
    )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Factory Data"

    # Get data rows
    rows = data.get('data', [])

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for the given filters")

    # Get column headers from first row
    headers = list(rows[0].keys())

    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"factory_data_{timestamp}.xlsx"

    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/apps/{app_name}/tables/application_status/data")
async def get_application_status_data(
    app_name: str = Path(..., description="Application name"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    yearmonth: Optional[str] = Query(None, description="Filter by YearMonth (format: 2026.01 or 2026-01), supports multiple values separated by comma"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get application status data with optional YearMonth filtering.

    This endpoint retrieves data from the application_status table and supports
    filtering by YearMonth field using Qlik selections. The YearMonth field
    exists in the app but may not be in the table itself.

    **Examples:**

    Get all data (no filtering):
    ```
    GET /api/v1/apps/Stock/tables/application_status/data?page=1&page_size=100
    ```

    Filter by single month:
    ```
    GET /api/v1/apps/Stock/tables/application_status/data?page=1&page_size=100&yearmonth=2024-01
    ```

    Filter by multiple months:
    ```
    GET /api/v1/apps/Stock/tables/application_status/data?page=1&page_size=100&yearmonth=2024-01,2024-02,2024-03
    ```

    **Response format:**
    ```json
    {
      "object_id": "UWDJj",
      "app_id": "...",
      "app_name": "Stock",
      "data": [
        {
          "Field1": "Value1",
          "Field2": "Value2",
          ...
        }
      ],
      "pagination": {
        "page": 1,
        "page_size": 100,
        "total_rows": 150,
        "total_pages": 2,
        "has_next": true,
        "has_previous": false
      }
    }
    ```
    """
    table_name = "application_status"

    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Get bookmark ID for this table
    # The bookmark pre-filters the data to 3 months for fast retrieval
    bookmark_id = settings.get_bookmark_id(app_name, table_name)

    # application_status uses the regular table object (UWDJj)
    # Apply bookmark FIRST to filter the data before creating the hypercube
    data = await app_service.get_object_data(
        app_name=app_name,
        object_id=object_id,
        page=page,
        page_size=page_size,
        filters={},  # No client-side filtering
        selections={},  # No Qlik selections - rely on bookmark
        variables={},  # No variables for this endpoint
        bookmark_id=bookmark_id  # Applied FIRST to filter the data
    )

    return data


@router.get("/apps/{app_name}/tables/{table_name}/data")
async def get_table_data_with_measures(
    app_name: str = Path(..., description="Application name"),
    table_name: str = Path(..., description="Table name (e.g., stock_qty)"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=10000, description="Rows per page"),
    factory: Optional[str] = Query(None, description="Filter by factory (PRCTR field)"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse (LGORT field)"),
    app_service: AppService = Depends(get_app_service),
    api_key: str = Depends(verify_api_key)
):
    """
    Get actual data from a table with dimensions and measures.

    This endpoint retrieves actual data rows where each row contains
    all dimension values and calculated measure values.

    **Example:**
    ```
    GET /api/v1/apps/Stock/tables/stock_qty/data?page=1&page_size=10
    ```

    **With filters:**
    ```
    GET /api/v1/apps/Stock/tables/stock_qty/data?page=1&page_size=10&factory=1203&warehouse=P210
    ```

    **Response format:**
    ```json
    {
      "data": [
        {
          "MATNR": "000000001000000000",
          "Название материалов": "Замес д. ПВХ профилей PE",
          "База Код": "1203",
          "Название завода": "Производство ПВХ профилей",
          "Склад": "P210",
          "СвобИспользЗапас": 234.22,
          "Базовая ЕИ": "KG"
        }
      ],
      "pagination": {...}
    }
    ```
    """
    # Check app access
    if not settings.can_access_app(api_key, app_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to app '{app_name}'"
        )

    # Check table access
    if not settings.can_access_table(api_key, app_name, table_name):
        raise HTTPException(
            status_code=403,
            detail=f"Your API key does not have access to table '{table_name}' in app '{app_name}'"
        )

    # Get object ID for this table
    object_id = settings.get_object_id_for_table(app_name, table_name)
    if not object_id:
        raise HTTPException(
            status_code=404,
            detail=f"No object mapping found for table '{table_name}' in app '{app_name}'"
        )

    # Build filters dictionary
    # Map query parameters to actual Qlik field names
    filters = {}
    if factory:
        filters['PRCTR'] = factory  # PRCTR is the factory field in Qlik
    if warehouse:
        filters['LGORT'] = warehouse  # LGORT is the warehouse field in Qlik

    data = await app_service.get_object_data(app_name, object_id, page, page_size, filters, selections={}, variables={})
    return data
